import streamlit as st
import pandas as pd
from collections import Counter
from datetime import date
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
import openpyxl
from io import BytesIO
import logging
import traceback

# =========================
# LOGGING
# =========================
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
)
logger = logging.getLogger("corpored")


st.set_page_config(page_title="Generador y Validador -- Ordenes Corpored", layout="centered")

# =========================
# ESTILOS
# =========================
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@400;500;600;700&family=DM+Mono:wght@400;500&display=swap');

html, body, [class*="css"] { font-family: 'DM Sans', sans-serif; }
.stApp { background-color: #f4f5f7; color: #1a1e2e; }

.header-block {
    background: #1a1e2e;
    border-left: 6px solid #e8a020;
    padding: 22px 30px; margin-bottom: 28px; border-radius: 0 8px 8px 0;
}
.header-block h1 {
    font-family: "DM Sans", sans-serif; font-size: 1.9rem; font-weight: 700;
    color: #ffffff; margin: 0; letter-spacing: 1px;
}
.header-block p {
    color: #e8a020; margin: 5px 0 0 0; font-size: 0.75rem;
    letter-spacing: 2px; text-transform: uppercase; font-family: "DM Mono", monospace;
}

.seccion-titulo {
    font-size: 0.7rem; font-weight: 700; letter-spacing: 2.5px; color: #1a1e2e;
    margin-bottom: 14px; text-transform: uppercase;
    border-bottom: 2px solid #e8a020; padding-bottom: 6px; display: inline-block;
}

.tabla-camion {
    width: 100%; border-collapse: collapse; font-size: 0.82rem; margin-top: 12px;
    background: #fff; border-radius: 6px; overflow: hidden;
    box-shadow: 0 1px 4px rgba(0,0,0,0.08);
}
.tabla-camion th {
    background: #1a1e2e; color: #e8a020; padding: 8px 12px; text-align: center;
    border: none; font-size: 0.7rem; letter-spacing: 1.5px; text-transform: uppercase;
}
.tabla-camion td {
    padding: 7px 12px; border-bottom: 1px solid #eef0f4; text-align: center;
    font-weight: 500; color: #1a1e2e; background: #fff;
}
.tabla-camion tr:last-child td { border-bottom: none; }

.preview-table {
    width: 100%; border-collapse: collapse; font-size: 0.82rem; margin-top: 8px;
    background: #fff; box-shadow: 0 1px 4px rgba(0,0,0,0.07);
    border-radius: 6px; overflow: hidden;
}
.preview-table th {
    background: #1a1e2e; color: #e8e8e8; padding: 8px 10px; border: none;
    text-align: center; font-size: 0.7rem; letter-spacing: 1px; text-transform: uppercase;
}
.preview-table td {
    padding: 6px 10px; border-bottom: 1px solid #eef0f4;
    text-align: center; color: #1a1e2e;
}
.preview-table tr:last-child td { border-bottom: none; }

.badge-ccr    { background:#dbeafe; color:#1d4ed8; padding:2px 9px; border-radius:12px; font-size:0.7rem; font-weight:600; }
.badge-fraval { background:#fee2e2; color:#b91c1c; padding:2px 9px; border-radius:12px; font-size:0.7rem; font-weight:600; }
.badge-auto   { background:#dcfce7; color:#15803d; padding:2px 9px; border-radius:12px; font-size:0.7rem; font-weight:600; }
.badge-manual { background:#fef9c3; color:#854d0e; padding:2px 9px; border-radius:12px; font-size:0.7rem; font-weight:600; }

div[data-testid="stNumberInput"] input,
div[data-testid="stTextInput"] input,
div[data-testid="stDateInput"] input {
    background: #ffffff !important; color: #1a1e2e !important;
    border: 1.5px solid #d1d5db !important; border-radius: 6px !important;
    font-family: "DM Mono", monospace !important; font-size: 0.85rem !important;
}
label { color: #374151 !important; font-size: 0.8rem !important; font-weight: 500 !important; }
.stButton > button {
    font-family: "DM Sans", sans-serif !important; font-weight: 600 !important;
    border-radius: 6px !important; font-size: 0.82rem !important;
}
hr { border-color: #e5e7eb !important; margin: 20px 0 !important; }
div[data-testid="metric-container"] {
    background: #fff; border: 1px solid #e5e7eb; border-radius: 8px; padding: 12px 16px;
}

/* Ocultar botones +/- del number_input del Destino individual (cambio funcional). */
[data-testid="stElementContainer"]:has(.no-stepper-anchor) + [data-testid="stElementContainer"] [data-testid="stNumberInputStepUp"],
[data-testid="stElementContainer"]:has(.no-stepper-anchor) + [data-testid="stElementContainer"] [data-testid="stNumberInputStepDown"],
div:has(> .no-stepper-anchor) + div [data-testid="stNumberInputStepUp"],
div:has(> .no-stepper-anchor) + div [data-testid="stNumberInputStepDown"],
div:has(> .no-stepper-anchor) ~ div [data-testid="stNumberInputStepUp"],
div:has(> .no-stepper-anchor) ~ div [data-testid="stNumberInputStepDown"] {
    display: none !important;
}
</style>
""", unsafe_allow_html=True)


# =========================
# CONSTANTES
# =========================
VEHICULO_TN     = "FZS4202"
VEHICULO_TC     = "FZS3567"
CLIENTE_CCR     = "0001500990"
CLIENTE_FRAVAL  = "0000203063"
LIMITE_FRAVAL   = 1400
COMERCIALIZADOR = "0001700001"
RFC_CHOFER      = "CAPM-750512-CU8"

# Comodines: el primer elemento es el default (igual a la constante de arriba)
VEHICULOS_TN_OPCIONES = ["FZS4202", "FZS4203", "FZS4204"]
VEHICULOS_TC_OPCIONES = ["FZS3567", "FZS-3731", "FZS-3933"]
RFC_CHOFER_OPCIONES   = ["CAPM-750512-CU8", "CURC-680915-SA5", "DUBR-780319-HV0"]
# Set para detectar "esta orden es de compartimentos" sin importar qué comodín se usó
VEHICULOS_COMPARTIMENTO = set(VEHICULOS_TC_OPCIONES)
CENTRO          = 703
MAPA_PRODUCTO   = {"MAGNA": 32011, "PREMIUM": 32012, "DIESEL": 34006}
# Pcarga por producto: MAGNA=1, DIESEL=3, PREMIUM=8
MAPA_PCARGA     = {"MAGNA": 1, "PREMIUM": 8, "DIESEL": 3}
EQUIVALENCIAS   = {"M": "MAGNA", "P": "PREMIUM", "D": "DIESEL"}
OPCIONES_PROD   = ["MAGNA", "PREMIUM", "DIESEL"]
# Para compartimentos: además permitimos N/A (compartimento sin orden)
NA_VALUE        = "N/A"
OPCIONES_PROD_COMP = OPCIONES_PROD + [NA_VALUE]
TONELES_TN_VALIDOS = {1, 2}                    # Tonel Normal solo permite 1 o 2
DESTINO_MAX     = 9999                          # Cota razonable de destino
MIME_XLSX       = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# Hojas y columnas esperadas en el Excel de entrada
HOJA_PROGRAMA       = "Programa del día "
HOJA_FIRME          = "FIRME"
HOJA_COMPARTIMENTOS = "Compartimentos"
# La clave de FIRME se construye en código a partir de:
COL_FIRME_DESTINO   = 2     # columna C — primeros 4 caracteres, solo dígitos
COL_FIRME_PRODUCTO  = 3     # columna D — MAGNA/PREMIUM/DIESEL
COL_FIRME_CAPACIDAD = 10    # columna K — "Capacidad programada (Litros)"
COL_FIRME_TONEL     = 9     # columna J — número de tonel (1, 2, 3, 4)
# Capacidad esperada por tonel para que una fila valide un compartimento:
# - Toneles 1 y 3 → 11,000 L
# - Toneles 2 y 4 → 20,000 L
LITROS_COMPARTIMENTO = {1: 11000, 2: 20000, 3: 11000, 4: 20000}
# Tonel Normal requiere ≥ 32,000 litros (cualquier valor por debajo invalida).
# Las filas con 11,000 o 20,000 NO validan Tonel Normal: solo compartimentos.
LITROS_TONEL_NORMAL_MIN = 32000
# Destinos fijos que se muestran en el resumen del Excel FIRME extra
DESTINOS_RESUMEN_FIRME_EXTRA = [87, 88, 89, 467]
COL_PROG_DEST_1     = 12    # M
COL_PROG_PROD_1     = 14    # O
COL_PROG_DEST_2     = 16    # Q
COL_PROG_PROD_2     = 18    # S


# =========================
# EXCEPCIONES PROPIAS
# =========================
class ValidationError(Exception):
    """Error de validación de entrada del usuario."""
    pass

class ExcelStructureError(Exception):
    """El Excel no tiene la estructura esperada."""
    pass


# =========================
# UTILIDADES
# =========================
def limpiar(valor):
    if pd.isna(valor):
        return ""
    valor = str(valor).strip().upper()
    if valor.endswith(".0"):
        valor = valor[:-2]
    return valor

def norm(p):
    p = str(p).strip().upper()
    return EQUIVALENCIAS.get(p, p)

def color_prod(p):
    return {"MAGNA": "#30B84C", "PREMIUM": "#E00D21", "DIESEL": "#211B1B"}.get(p, "#333")

def color_text(p):
    return {"MAGNA": "#000", "PREMIUM": "#fff", "DIESEL": "#fff"}.get(p, "#fff")


def render_resumen_firme_extra_html(df):
    """
    Construye el HTML de la matriz resumen para mostrar en pantalla.
    Mismo formato que el bloque resumen del Excel:
      - Filas: MAGNA / PREMIUM / DIESEL (con sus colores)
      - Columnas: destinos fijos × Tonel 1 / Tonel 2
      - Celdas: conteo de órdenes (vacío si 0)
    """
    destinos  = DESTINOS_RESUMEN_FIRME_EXTRA
    productos = ["MAGNA", "PREMIUM", "DIESEL"]

    # Conteos (solo toneles 1 y 2 y solo los destinos fijos)
    conteo = Counter()
    for _, r in df.iterrows():
        try:
            d = int(r["#"])
            t = int(r["Numero de Tonel"])
        except (ValueError, TypeError):
            continue
        p = r["Nombre Prod"]
        if t in (1, 2) and p in MAPA_PRODUCTO and d in destinos:
            conteo[(d, p, t)] += 1

    # Si no hay ningún conteo en los destinos fijos, devolver None
    if sum(conteo.values()) == 0:
        return None

    # Construcción del HTML
    th_style = ('background:#1a1e2e;color:#e8a020;padding:8px 12px;text-align:center;'
                'font-size:0.72rem;letter-spacing:1px;text-transform:uppercase;font-weight:700;'
                'border:1px solid #1a1e2e;')
    td_style = ('padding:7px 12px;border:1px solid #eef0f4;text-align:center;'
                'font-weight:600;color:#1a1e2e;background:#fff;')
    total_label_style = ('background:#1a1e2e;color:#e8a020;padding:8px 12px;text-align:center;'
                         'font-size:0.75rem;letter-spacing:1px;text-transform:uppercase;'
                         'font-weight:700;border:1px solid #1a1e2e;')
    total_cell_style = ('padding:8px 12px;border:1px solid #1a1e2e;text-align:center;'
                        'font-weight:700;color:#1a1e2e;background:#f5f593;font-size:0.9rem;')

    # Fila 1: destinos (cada uno con colspan=2)
    fila_destinos = f'<th style="{th_style}">Destino</th>'
    for d in destinos:
        fila_destinos += f'<th colspan="2" style="{th_style}">{d}</th>'

    # Fila 2: Producto | Tonel 1 | Tonel 2 | ...
    fila_toneles = f'<th style="{th_style}">Producto</th>'
    for _ in destinos:
        fila_toneles += f'<th style="{th_style}">Tonel 1</th><th style="{th_style}">Tonel 2</th>'

    # Filas de productos
    filas_body = ""
    for prod in productos:
        bg = color_prod(prod); fg = color_text(prod)
        prod_cell = (f'<td style="background:{bg};color:{fg};padding:7px 12px;'
                     f'border:1px solid #eef0f4;text-align:center;font-weight:700;">{prod}</td>')
        body = prod_cell
        for d in destinos:
            for t in (1, 2):
                count = conteo.get((d, prod, t), 0)
                body += f'<td style="{td_style}">{count if count else ""}</td>'
        filas_body += f"<tr>{body}</tr>"

    # Fila final: "Total" + totales por destino (T1+T2 mergeados) + total general
    total_general = sum(conteo.values())
    fila_total = f'<td style="{total_label_style}">Total</td>'
    for d in destinos:
        total_dest = sum(conteo.get((d, p, t), 0) for p in productos for t in (1, 2))
        fila_total += f'<td colspan="2" style="{total_cell_style}">{total_dest if total_dest else ""}</td>'

    # Etiqueta de total general arriba a la derecha + valor en la fila final
    # Para que se vea bien sin desbalancear columnas, lo mostramos como una pequeña
    # leyenda fuera de la tabla.
    leyenda_total = (
        f'<div style="text-align:right;margin-top:6px;font-size:0.85rem;'
        f'color:#1a1e2e;font-weight:700;">'
        f'<span style="background:#1a1e2e;color:#e8a020;padding:6px 14px;'
        f'border-radius:4px;letter-spacing:1px;text-transform:uppercase;'
        f'font-size:0.75rem;">Total general:</span> '
        f'<span style="background:#f5f593;padding:6px 14px;border-radius:4px;'
        f'border:2px solid #1a1e2e;">{total_general}</span></div>'
    )

    html = f"""<table style="width:100%;border-collapse:collapse;font-size:0.82rem;
        background:#fff;border-radius:6px;overflow:hidden;
        box-shadow:0 1px 4px rgba(0,0,0,0.08);margin-top:8px;">
        <tr>{fila_destinos}</tr>
        <tr>{fila_toneles}</tr>
        {filas_body}
        <tr>{fila_total}</tr>
    </table>
    {leyenda_total}"""
    return html


def _parsear_capacidad(l_raw):
    """
    Parsea el valor de la columna K (capacidad). Acepta enteros, floats,
    strings con comas/espacios. Retorna int o None si no se puede parsear.
    """
    if pd.isna(l_raw):
        return None
    try:
        return int(float(str(l_raw).strip().replace(",", "")))
    except (ValueError, TypeError):
        return None


def _clave_base(row):
    """
    Construye solo la parte destino+producto+tonel de la clave de FIRME
    (sin validar capacidad). Retorna (clave, tonel) o (None, None) si los
    datos básicos están incompletos.
    """
    # Columna C — destino
    c_raw = row.iloc[COL_FIRME_DESTINO]
    if pd.isna(c_raw):
        return None, None
    c_str = str(c_raw).strip()
    if not c_str:
        return None, None
    primeros_4 = c_str[:4]
    destino = "".join(ch for ch in primeros_4 if ch.isdigit())
    if not destino:
        return None, None

    # Columna D — producto
    d_raw = row.iloc[COL_FIRME_PRODUCTO]
    if pd.isna(d_raw):
        return None, None
    d_upper = str(d_raw).strip().upper()
    producto = None
    for nombre in MAPA_PRODUCTO:        # MAGNA, PREMIUM, DIESEL
        if nombre in d_upper:
            producto = nombre
            break
    if producto is None:
        return None, None

    # Columna J — tonel
    j_raw = row.iloc[COL_FIRME_TONEL]
    if pd.isna(j_raw):
        return None, None
    try:
        tonel = int(float(str(j_raw).strip()))
    except (ValueError, TypeError):
        return None, None
    if tonel <= 0:
        return None, None

    return f"{destino}{producto}{tonel}", tonel


def construir_clave_firme(row):
    """
    Construye la clave de FIRME para validar TONEL NORMAL.
    Requisitos:
      - Datos básicos válidos (C, D, J).
      - Tonel ∈ {1, 2} (los toneles 3 y 4 son siempre compartimentos).
      - Capacidad en columna K ≥ 32,000 litros.
        (Filas con 11,000 o 20,000 son compartimentos y NO validan Tonel Normal.)

    Retorna la clave (ej: "87MAGNA1") o None si la fila no aplica.
    """
    clave, tonel = _clave_base(row)
    if clave is None:
        return None

    # Defensa extra: Tonel Normal solo existe con tonel 1 o 2
    if tonel not in TONELES_TN_VALIDOS:
        return None

    capacidad = _parsear_capacidad(row.iloc[COL_FIRME_CAPACIDAD])
    if capacidad is None:
        return None
    if capacidad < LITROS_TONEL_NORMAL_MIN:
        return None

    return clave


def clave_firme_para_compartimentos(row):
    """
    Construye la clave SOLO si la fila también cumple el requisito de litros
    para validar un compartimento:
      - Toneles 1 y 3 → 11,000 L exactos en columna K
      - Toneles 2 y 4 → 20,000 L exactos en columna K
    Retorna la clave (ej: "87MAGNA1") o None si la fila no es válida para
    validar un compartimento (faltan datos, capacidad incorrecta, o tonel
    fuera del rango 1-4).
    """
    clave, tonel = _clave_base(row)
    if clave is None:
        return None

    esperado = LITROS_COMPARTIMENTO.get(tonel)
    if esperado is None:
        # Tonel fuera de 1-4: no cuenta como compartimento
        return None

    capacidad = _parsear_capacidad(row.iloc[COL_FIRME_CAPACIDAD])
    if capacidad is None:
        return None
    if capacidad != esperado:
        return None

    return clave


# =========================
# VALIDACIONES REUTILIZABLES
# =========================
def validar_destino(raw, etiqueta="Destino"):
    """
    Valida que `raw` sea un entero positivo válido como destino.
    Retorna el int. Lanza ValidationError con mensaje claro si falla.
    """
    if raw is None:
        raise ValidationError(f"{etiqueta} vacío")
    s = str(raw).strip()
    if not s:
        raise ValidationError(f"{etiqueta} vacío")
    if "." in s:
        raise ValidationError(f"{etiqueta} '{s}' no válido — debe ser entero (sin decimales)")
    if not s.lstrip("-").isdigit():
        raise ValidationError(f"{etiqueta} '{s}' no válido — solo se aceptan números enteros")
    valor = int(s)
    if valor <= 0:
        raise ValidationError(f"{etiqueta} '{s}' no válido — debe ser un número positivo")
    if valor > DESTINO_MAX:
        raise ValidationError(f"{etiqueta} '{s}' fuera de rango (máx {DESTINO_MAX})")
    return valor

def validar_tonel_tn(raw, etiqueta="Tonel"):
    """
    Valida tonel para Tonel Normal: solo 1 o 2 son aceptados.
    Retorna el int. Lanza ValidationError si falla.
    """
    if raw is None:
        raise ValidationError(f"{etiqueta} vacío")
    s = str(raw).strip()
    if not s:
        raise ValidationError(f"{etiqueta} vacío")
    if not s.isdigit():
        raise ValidationError(f"{etiqueta} '{s}' no válido — debe ser 1 o 2")
    valor = int(s)
    if valor not in TONELES_TN_VALIDOS:
        raise ValidationError(f"{etiqueta} '{s}' no válido — solo se aceptan los valores {sorted(TONELES_TN_VALIDOS)}")
    return valor

def validar_producto(raw, etiqueta="Producto"):
    """
    Valida que el producto sea M/P/D o MAGNA/PREMIUM/DIESEL.
    Retorna el nombre canónico. Lanza ValidationError si falla.
    """
    if raw is None:
        raise ValidationError(f"{etiqueta} vacío")
    s = str(raw).strip()
    if not s:
        raise ValidationError(f"{etiqueta} vacío")
    p = norm(s)
    if p not in MAPA_PRODUCTO:
        validos = ", ".join(MAPA_PRODUCTO.keys())
        raise ValidationError(f"{etiqueta} '{s}' desconocido — válidos: {validos} (o M/P/D)")
    return p


# =========================
# RECOLECTAR REGISTROS MANUALES
# =========================
def recolectar_manuales():
    frames = []
    if st.session_state.get("tn_individual"):
        frames.append(pd.DataFrame(st.session_state.tn_individual))
    if st.session_state.get("tn_masivo"):
        frames.append(pd.DataFrame(st.session_state.tn_masivo))
    if st.session_state.get("compartimentos_tc"):
        # Usar la selección actual de unidad y operador para compartimentos.
        # Esto hace que cambiar el selector se refleje automáticamente.
        vehiculo_actual = st.session_state.get("vehiculo_tc", VEHICULO_TC)
        rfc_actual      = st.session_state.get("rfc_tc", RFC_CHOFER)
        registros = []
        for cam in st.session_state.compartimentos_tc:
            for prod, tonel in zip(cam["prods"], [1, 2, 3, 4]):
                # Omitir compartimentos marcados como N/A
                if prod == NA_VALUE:
                    continue
                destino = cam["d1"] if tonel <= 2 else cam["d2"]
                registros.append({
                    "#": destino, "Nombre Prod": prod,
                    "Clve de producto": MAPA_PRODUCTO[prod],
                    "Numero de Tonel": tonel,
                    "Clave de Vehiculo": vehiculo_actual,
                    "R.F.C. chofer":     rfc_actual,
                    "Tipo": "COMPARTIMENTO", "origen": "manual"
                })
        if registros:
            frames.append(pd.DataFrame(registros))
    return pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()


# =========================
# SESSION STATE INICIAL
# =========================
for key, default in [
    ("tn_individual", []),
    ("tn_masivo", []),
    ("compartimentos_tc", []),
    ("vehiculo_tn",   VEHICULOS_TN_OPCIONES[0]),
    ("rfc_tn",        RFC_CHOFER_OPCIONES[0]),
    ("vehiculo_tc",   VEHICULOS_TC_OPCIONES[0]),
    ("rfc_tc",        RFC_CHOFER_OPCIONES[0]),
]:
    if key not in st.session_state:
        st.session_state[key] = default


# =========================
# PROCESAR ARCHIVO
# =========================
def procesar(archivo):
    """
    Procesa el Excel y devuelve DF de faltantes.
    Lanza ExcelStructureError si el archivo no tiene la estructura esperada.
    """
    # Verificar hojas esperadas antes de leer
    try:
        xls = pd.ExcelFile(archivo)
    except Exception as e:
        raise ExcelStructureError(f"No se pudo abrir el archivo Excel: {e}") from e

    hojas_requeridas = [HOJA_PROGRAMA, HOJA_FIRME, HOJA_COMPARTIMENTOS]
    faltan = [h for h in hojas_requeridas if h not in xls.sheet_names]
    if faltan:
        raise ExcelStructureError(
            f"El archivo no contiene las hojas requeridas: {faltan}. "
            f"Hojas encontradas: {xls.sheet_names}"
        )

    try:
        df_prog  = pd.read_excel(archivo, sheet_name=HOJA_PROGRAMA)
        df_firme = pd.read_excel(archivo, sheet_name=HOJA_FIRME)
    except Exception as e:
        raise ExcelStructureError(f"Error leyendo hojas del Excel: {e}") from e

    # Validar que las hojas tengan suficientes columnas
    cols_min_prog = max(COL_PROG_DEST_1, COL_PROG_PROD_1, COL_PROG_DEST_2, COL_PROG_PROD_2) + 1
    if df_prog.shape[1] < cols_min_prog:
        raise ExcelStructureError(
            f"La hoja '{HOJA_PROGRAMA}' tiene {df_prog.shape[1]} columnas, "
            f"se esperaban al menos {cols_min_prog}."
        )
    cols_min_firme = max(COL_FIRME_DESTINO, COL_FIRME_PRODUCTO,
                         COL_FIRME_CAPACIDAD, COL_FIRME_TONEL) + 1
    if df_firme.shape[1] < cols_min_firme:
        raise ExcelStructureError(
            f"La hoja '{HOJA_FIRME}' tiene {df_firme.shape[1]} columnas, "
            f"se esperaban al menos {cols_min_firme} (C, D, J y K)."
        )

    try:
        wb        = load_workbook(archivo, data_only=True)
        hoja_comp = wb[HOJA_COMPARTIMENTOS]
    except Exception as e:
        raise ExcelStructureError(f"Error abriendo hoja '{HOJA_COMPARTIMENTOS}': {e}") from e

    faltantes = []

    # ── CLAVES FIRME ──
    # Antes: se leían de las columnas R y S (fórmulas en Excel).
    # Ahora: se construyen en código.
    # Dos cuentas distintas:
    #   - conteo_firme_prog: para validar Tonel Normal. Solo requiere C+D+J.
    #   - conteo_firme_comp: para validar Compartimentos. Además exige que
    #     la columna K (Capacidad programada) coincida exactamente con los
    #     litros esperados para ese tonel (11,000 para 1/3, 20,000 para 2/4).
    claves_firme_prog = [
        clave for clave in (construir_clave_firme(row) for _, row in df_firme.iterrows())
        if clave is not None
    ]
    conteo_firme_prog = Counter(claves_firme_prog)

    claves_firme_comp = [
        clave for clave in (clave_firme_para_compartimentos(row) for _, row in df_firme.iterrows())
        if clave is not None
    ]
    conteo_firme_comp = Counter(claves_firme_comp)

    # ── PROGRAMA ──
    concatenaciones = []
    for _, row in df_prog.iterrows():
        d1 = limpiar(row.iloc[COL_PROG_DEST_1])
        p1 = limpiar(row.iloc[COL_PROG_PROD_1])
        if d1 and p1 and "/" not in p1 and p1 in MAPA_PRODUCTO:
            concatenaciones.append((d1, p1, "1"))

        d2 = limpiar(row.iloc[COL_PROG_DEST_2])
        p2 = limpiar(row.iloc[COL_PROG_PROD_2])
        if d2 and p2 and "/" not in p2 and p2 in MAPA_PRODUCTO:
            concatenaciones.append((d2, p2, "2"))

    conteo_prog    = Counter(d + p + t for d, p, t in concatenaciones)
    conteo_ya_prog = Counter()

    for d, p, t in concatenaciones:
        clave = d + p + t
        total = conteo_prog[clave] - conteo_firme_prog.get(clave, 0)
        if total > 0 and conteo_ya_prog[clave] < total:
            faltantes.append((d, p, t, "", VEHICULO_TN, "NORMAL"))
            conteo_ya_prog[clave] += 1

    # ── COMPARTIMENTOS ──
    # Usa `conteo_firme_comp`, que solo cuenta filas de FIRME cuyos litros
    # coinciden con el tonel (validación reforzada).
    lista             = []

    for fila in range(2, 11):
        m = limpiar(hoja_comp[f"M{fila}"].value)
        q = limpiar(hoja_comp[f"Q{fila}"].value)
        n = limpiar(hoja_comp[f"N{fila}"].value)
        o = limpiar(hoja_comp[f"O{fila}"].value)
        r = limpiar(hoja_comp[f"R{fila}"].value)
        s = limpiar(hoja_comp[f"S{fila}"].value)

        for a, b, c in [(m, n, "1"), (m, o, "2"), (q, r, "3"), (q, s, "4")]:
            if a and b and "/" not in b:
                lista.append((a, b, c, a + b + c))

    conteo_comp    = Counter(x[3] for x in lista)
    conteo_ya_comp = Counter()

    for a, b, c, clave in lista:
        total = conteo_comp[clave] - conteo_firme_comp.get(clave, 0)
        if total > 0 and conteo_ya_comp[clave] < total:
            faltantes.append((a, b, c, c, VEHICULO_TC, "COMPARTIMENTO"))
            conteo_ya_comp[clave] += 1

    # ── FIRME EXTRA: filas de FIRME que NO están en Programa+Compartimentos ──
    # Para Tonel Normal: keys del Programa son `d + p + t` (con t ∈ {1,2}).
    # Para Compartimentos: keys son `a + b + c` (con c ∈ {1,2,3,4}).
    # Una fila de FIRME es "extra" si su clave aparece más veces en FIRME
    # que en el origen correspondiente (Programa o Compartimentos).
    firme_extra = []

    # Re-iterar FIRME para extraer la información completa de cada fila válida.
    # Tonel Normal (FIRME válido para TN) → comparar contra conteo_prog
    conteo_consumido_prog = Counter()
    for _, row in df_firme.iterrows():
        clave = construir_clave_firme(row)
        if clave is None:
            continue
        # ¿Cuántos de esta clave hay en el programa?
        en_programa = conteo_prog.get(clave, 0)
        if conteo_consumido_prog[clave] < en_programa:
            conteo_consumido_prog[clave] += 1
            continue  # esta fila de FIRME está respaldada por el programa
        # No respaldada → es FIRME extra
        # Reconstruir destino, producto, tonel
        c_str = str(row.iloc[COL_FIRME_DESTINO]).strip()
        destino = "".join(ch for ch in c_str[:4] if ch.isdigit())
        d_upper = str(row.iloc[COL_FIRME_PRODUCTO]).strip().upper()
        producto = next((n for n in MAPA_PRODUCTO if n in d_upper), None)
        tonel = int(float(str(row.iloc[COL_FIRME_TONEL]).strip()))
        firme_extra.append((destino, producto, str(tonel), "", VEHICULO_TN, "NORMAL"))

    # Compartimentos (FIRME válido para Comp) → comparar contra conteo_comp
    conteo_consumido_comp = Counter()
    for _, row in df_firme.iterrows():
        clave = clave_firme_para_compartimentos(row)
        if clave is None:
            continue
        en_comp = conteo_comp.get(clave, 0)
        if conteo_consumido_comp[clave] < en_comp:
            conteo_consumido_comp[clave] += 1
            continue
        # FIRME extra de compartimento
        c_str = str(row.iloc[COL_FIRME_DESTINO]).strip()
        destino = "".join(ch for ch in c_str[:4] if ch.isdigit())
        d_upper = str(row.iloc[COL_FIRME_PRODUCTO]).strip().upper()
        producto = next((n for n in MAPA_PRODUCTO if n in d_upper), None)
        tonel = int(float(str(row.iloc[COL_FIRME_TONEL]).strip()))
        firme_extra.append((destino, producto, str(tonel), str(tonel), VEHICULO_TC, "COMPARTIMENTO"))

    df_extra = pd.DataFrame(firme_extra, columns=[
        "#", "Nombre Prod", "Numero de Tonel", "Tipo interno", "Clave de Vehiculo", "Tipo"
    ])
    if not df_extra.empty:
        df_extra["Clve de producto"] = df_extra["Nombre Prod"].map(MAPA_PRODUCTO)
        df_extra = df_extra[(df_extra["Numero de Tonel"] != "0") & (df_extra["Clve de producto"].notna())]
        df_extra["#"] = pd.to_numeric(df_extra["#"], errors="coerce")
        df_extra = df_extra.dropna(subset=["#"])
        df_extra["#"] = df_extra["#"].astype(int)
        df_extra = df_extra.drop(columns=["Tipo interno"])
        df_extra["origen"] = "firme_extra"
    else:
        df_extra = None

    df = pd.DataFrame(faltantes, columns=[
        "#", "Nombre Prod", "Numero de Tonel", "Tipo interno", "Clave de Vehiculo", "Tipo"
    ])

    if df.empty:
        return None, df_extra

    df["Clve de producto"] = df["Nombre Prod"].map(MAPA_PRODUCTO)
    df = df[(df["Numero de Tonel"] != "0") & (df["Clve de producto"].notna())]
    df["#"] = pd.to_numeric(df["#"], errors="coerce")
    df = df.dropna(subset=["#"])
    df["#"] = df["#"].astype(int)
    df = df.drop(columns=["Tipo interno"])
    df["origen"] = "auto"
    return df, df_extra


# =========================
# CONSTRUIR FORMATO SIIC
# =========================
def construir(df, fecha_usuario, cliente):
    df = df.copy()
    df["Numero de Tonel"]     = pd.to_numeric(df["Numero de Tonel"], errors="coerce")
    df["Centro"]              = CENTRO
    df["Fecha p"]             = fecha_usuario
    df["Pedido"]              = ""
    df["Comercializador"]     = COMERCIALIZADOR
    df["Tipo movimiento"]     = "SV"
    df["Clve de cliente"]     = cliente
    df["Presentación"]        = 3
    df["Formula"]             = 2
    df["Turno"]               = 1
    df["Medio de transporte"] = "AT"
    df["Espacio en blanco"]   = ""
    # Si la orden ya trae R.F.C. chofer (manual), respetarlo; si no, default.
    if "R.F.C. chofer" not in df.columns:
        df["R.F.C. chofer"] = RFC_CHOFER
    else:
        df["R.F.C. chofer"] = df["R.F.C. chofer"].fillna(RFC_CHOFER).replace("", RFC_CHOFER)
    # Pcarga depende del producto: MAGNA=1, DIESEL=3, PREMIUM=8.
    # Si el producto no se reconoce, usar 1 como fallback seguro.
    df["Pcarga"] = df["Nombre Prod"].map(MAPA_PCARGA).fillna(1).astype(int)
    df["Observaciones"]       = 2
    return df[[
        "Centro", "Fecha p", "Pedido", "Comercializador", "Tipo movimiento",
        "Clve de cliente", "#", "Clve de producto", "Nombre Prod",
        "Presentación", "Formula", "Turno", "Medio de transporte",
        "Clave de Vehiculo", "Numero de Tonel", "Espacio en blanco",
        "R.F.C. chofer", "Pcarga", "Observaciones"
    ]]


# =========================
# GENERAR EXCEL
# =========================
def generar_excel(df):
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Formato")
        ws = writer.book["Formato"]

        fills = {
            "header":  PatternFill(start_color="9E9DA3", fill_type="solid"),
            "default": PatternFill(start_color="F5F593", fill_type="solid"),
            "magna":   PatternFill(start_color="30B84C", fill_type="solid"),
            "premium": PatternFill(start_color="E00D21", fill_type="solid"),
            "diesel":  PatternFill(start_color="211B1B", fill_type="solid"),
            "comp_vehiculo": PatternFill(start_color="FAD225", fill_type="solid"),
        }
        bold   = Font(bold=True)
        white  = Font(bold=True, color="FFFFFF")
        border = Border(left=Side(style="thin"), right=Side(style="thin"),
                        top=Side(style="thin"),  bottom=Side(style="thin"))
        center = Alignment(horizontal="center", vertical="center")

        for c in ws[1]:
            c.fill = fills["header"]; c.font = bold
            c.border = border; c.alignment = center

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.fill = fills["default"]; cell.font = bold; cell.border = border
                if cell.value in ("MAGNA",   32011): cell.fill = fills["magna"]
                if cell.value in ("PREMIUM", 32012): cell.fill = fills["premium"]
                if cell.value in ("DIESEL",  34006):
                    cell.fill = fills["diesel"]; cell.font = white
                # Clave de Vehiculo (col N) cuando es vehículo de compartimentos
                if cell.column_letter == "N" and cell.value in VEHICULOS_COMPARTIMENTO:
                    cell.fill = fills["comp_vehiculo"]
                if cell.column_letter == "O":
                    # Conversión segura: solo intenta cuando el valor es convertible
                    if cell.value is not None and not isinstance(cell.value, (int, float)):
                        try:
                            cell.value = float(cell.value)
                        except (ValueError, TypeError) as e:
                            logger.warning(
                                "No se pudo convertir a float la celda O%s con valor %r: %s",
                                cell.row, cell.value, e
                            )
                    cell.number_format = "0"; cell.alignment = center
                elif isinstance(cell.value, (int, float)):
                    cell.alignment = center

        for col in ws.columns:
            max_len = max(len(str(c.value)) if c.value else 0 for c in col)
            ws.column_dimensions[col[0].column_letter].width = max_len + 2

    return output.getvalue()


def generar_excel_firme_extra(df, fecha_usuario):
    """
    Genera el Excel del tab FIRME extra con DOS bloques:

    1) BLOQUE RESUMEN (filas 1-5): matriz cruzada
       - Filas: MAGNA, PREMIUM, DIESEL
       - Columnas: destinos fijos (DESTINOS_RESUMEN_FIRME_EXTRA) × Tonel 1/Tonel 2
       - Valor: conteo de órdenes para cada (destino, producto, tonel)

    2) BLOQUE DETALLADO (después de una fila en blanco):
       6 columnas: Fecha, #, Clve de producto, Nombre Prod,
       Tipo de orden (Tonel Normal / Compartimento), Numero de Tonel.

    `df` debe tener: '#', 'Clve de producto', 'Nombre Prod',
    'Clave de Vehiculo', 'Numero de Tonel'.
    """
    # Fecha: 'ddmmaaaa' → 'dd-mm-aaaa'
    if len(fecha_usuario) == 8 and fecha_usuario.isdigit():
        fecha_fmt = f"{fecha_usuario[0:2]}-{fecha_usuario[2:4]}-{fecha_usuario[4:8]}"
    else:
        fecha_fmt = fecha_usuario

    def _tipo_orden(vehiculo):
        return "Compartimento" if vehiculo in VEHICULOS_COMPARTIMENTO else "Tonel Normal"

    # ── Pre-cálculo: conteos para el bloque resumen ──
    # Llave: (destino, producto, tonel) → cantidad
    conteo_resumen = Counter()
    for _, r in df.iterrows():
        try:
            d  = int(r["#"])
            t  = int(r["Numero de Tonel"])
        except (ValueError, TypeError):
            continue
        p = r["Nombre Prod"]
        if t in (1, 2) and p in MAPA_PRODUCTO:
            conteo_resumen[(d, p, t)] += 1

    # ── Estilos compartidos ──
    fills = {
        "header":  PatternFill(start_color="9E9DA3", fill_type="solid"),
        "default": PatternFill(start_color="F5F593", fill_type="solid"),
        "magna":   PatternFill(start_color="30B84C", fill_type="solid"),
        "premium": PatternFill(start_color="E00D21", fill_type="solid"),
        "diesel":  PatternFill(start_color="211B1B", fill_type="solid"),
        "tn":      PatternFill(start_color="FFFFFF", fill_type="solid"),
        "comp":    PatternFill(start_color="FAD225", fill_type="solid"),
        "white":   PatternFill(start_color="FFFFFF", fill_type="solid"),
    }
    bold   = Font(bold=True)
    white  = Font(bold=True, color="FFFFFF")
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"),  bottom=Side(style="thin"))
    center = Alignment(horizontal="center", vertical="center")

    fill_por_producto = {
        "MAGNA":   (fills["magna"],   bold),
        "PREMIUM": (fills["premium"], bold),
        "DIESEL":  (fills["diesel"],  white),
    }

    # ── Construcción manual con openpyxl ──
    output = BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ordenes_Extra"

    destinos    = DESTINOS_RESUMEN_FIRME_EXTRA
    n_destinos  = len(destinos)
    productos   = ["MAGNA", "PREMIUM", "DIESEL"]

    # Fila 1: encabezado de destinos
    # Col A: "Destino" + label "Producto" debajo. Col B+: destinos, 2 cols cada uno
    ws.cell(row=1, column=1, value="Destino").fill = fills["header"]
    ws.cell(row=1, column=1).font   = bold
    ws.cell(row=1, column=1).border = border
    ws.cell(row=1, column=1).alignment = center
    for i, d in enumerate(destinos):
        col_inicio = 2 + i * 2
        col_fin    = col_inicio + 1
        ws.cell(row=1, column=col_inicio, value=d)
        ws.merge_cells(start_row=1, start_column=col_inicio,
                       end_row=1,   end_column=col_fin)
        for c in (col_inicio, col_fin):
            cell = ws.cell(row=1, column=c)
            cell.fill = fills["header"]; cell.font = bold
            cell.border = border; cell.alignment = center

    # Fila 2: "Producto" + "Tonel 1 / Tonel 2" por destino
    ws.cell(row=2, column=1, value="Producto")
    ws.cell(row=2, column=1).fill = fills["header"]
    ws.cell(row=2, column=1).font = bold
    ws.cell(row=2, column=1).border = border
    ws.cell(row=2, column=1).alignment = center
    for i in range(n_destinos):
        col_t1 = 2 + i * 2
        col_t2 = col_t1 + 1
        ws.cell(row=2, column=col_t1, value="Tonel 1")
        ws.cell(row=2, column=col_t2, value="Tonel 2")
        for c in (col_t1, col_t2):
            cell = ws.cell(row=2, column=c)
            cell.fill = fills["header"]; cell.font = bold
            cell.border = border; cell.alignment = center

    # Filas 3-5: filas por producto
    for j, prod in enumerate(productos):
        fila = 3 + j
        prod_fill, prod_font = fill_por_producto[prod]
        # Columna A: nombre del producto con su color
        cell_a = ws.cell(row=fila, column=1, value=prod)
        cell_a.fill = prod_fill
        cell_a.font = prod_font
        cell_a.border = border
        cell_a.alignment = center
        # Una columna por (destino, tonel)
        for i, d in enumerate(destinos):
            for t_idx, t in enumerate((1, 2)):
                col = 2 + i * 2 + t_idx
                count = conteo_resumen.get((d, prod, t), 0)
                cell = ws.cell(row=fila, column=col, value=count if count else "")
                cell.fill = fills["default"]
                cell.font = bold
                cell.border = border
                cell.alignment = center

    # Fila 6: totales por destino (mergeando T1+T2 en una sola celda)
    cell_label = ws.cell(row=6, column=1, value="Total")
    cell_label.fill = fills["header"]
    cell_label.font = bold
    cell_label.border = border
    cell_label.alignment = center
    for i, d in enumerate(destinos):
        col_inicio = 2 + i * 2
        col_fin    = col_inicio + 1
        total_dest = sum(conteo_resumen.get((d, p, t), 0)
                         for p in productos for t in (1, 2))
        ws.cell(row=6, column=col_inicio, value=total_dest if total_dest else "")
        ws.merge_cells(start_row=6, start_column=col_inicio,
                       end_row=6,   end_column=col_fin)
        for c in (col_inicio, col_fin):
            cell = ws.cell(row=6, column=c)
            cell.fill = fills["default"]
            cell.font = bold
            cell.border = border
            cell.alignment = center

    # Fila 7: total general (en columna A con etiqueta, y celda mergeada con el valor)
    total_general = sum(conteo_resumen.values())
    cell_gtl = ws.cell(row=7, column=1, value="Total general")
    cell_gtl.fill = fills["header"]
    cell_gtl.font = bold
    cell_gtl.border = border
    cell_gtl.alignment = center
    # Valor de total general ocupa todas las columnas de destinos
    col_total_value_start = 2
    col_total_value_end   = 2 + len(destinos) * 2 - 1
    ws.cell(row=7, column=col_total_value_start, value=total_general)
    ws.merge_cells(start_row=7, start_column=col_total_value_start,
                   end_row=7,   end_column=col_total_value_end)
    for c in range(col_total_value_start, col_total_value_end + 1):
        cell = ws.cell(row=7, column=c)
        cell.fill = fills["default"]
        cell.font = bold
        cell.border = border
        cell.alignment = center

    # ── Bloque detallado (después de una fila en blanco) ──
    fila_inicio_detalle = 9  # 7 filas resumen (incl. totales) + 1 vacía + 1 header

    # Headers del detalle
    detalle_headers = ["Fecha", "#", "Clve de producto", "Nombre Prod",
                       "Tipo de orden", "Numero de Tonel"]
    for c, h in enumerate(detalle_headers, start=1):
        cell = ws.cell(row=fila_inicio_detalle, column=c, value=h)
        cell.fill = fills["header"]; cell.font = bold
        cell.border = border; cell.alignment = center

    # Filtrar: omitir las filas que ya están reflejadas en el resumen
    # (destinos fijos + tonel 1 o 2). Compartimentos a esos destinos sí entran.
    df_detalle = df[~(
        df["#"].isin(DESTINOS_RESUMEN_FIRME_EXTRA) &
        pd.to_numeric(df["Numero de Tonel"], errors="coerce").isin([1, 2])
    )]

    # Filas del detalle
    for i, (_, r) in enumerate(df_detalle.iterrows(), start=1):
        fila = fila_inicio_detalle + i
        try:
            destino_val = int(r["#"])
        except (ValueError, TypeError):
            destino_val = r["#"]
        try:
            tonel_val = int(r["Numero de Tonel"])
        except (ValueError, TypeError):
            tonel_val = r["Numero de Tonel"]

        valores = [
            fecha_fmt,
            destino_val,
            r["Clve de producto"],
            r["Nombre Prod"],
            _tipo_orden(r["Clave de Vehiculo"]),
            tonel_val,
        ]
        for c, v in enumerate(valores, start=1):
            cell = ws.cell(row=fila, column=c, value=v)
            cell.fill = fills["default"]; cell.font = bold
            cell.border = border; cell.alignment = center

            # Color por producto en columnas C y D
            if v in ("MAGNA",   32011): cell.fill = fills["magna"]
            if v in ("PREMIUM", 32012): cell.fill = fills["premium"]
            if v in ("DIESEL",  34006):
                cell.fill = fills["diesel"]; cell.font = white

            # Tipo de orden (col E): amarillo para Compartimento
            if c == 5:
                if v == "Compartimento":
                    cell.fill = fills["comp"]
                elif v == "Tonel Normal":
                    cell.fill = fills["tn"]

    # Anchos automáticos
    for col_idx in range(1, ws.max_column + 1):
        letter = openpyxl.utils.get_column_letter(col_idx)
        max_len = 0
        for cell in ws[letter]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = max(max_len + 2, 10)

    wb.save(output)
    return output.getvalue()


def aplicar_sufijo(nombre, sufijo):
    """
    Inserta un sufijo opcional entre el nombre y la extensión.
    Ej: ('SIIC_CCR_01012026.xlsx', 'v2') -> 'SIIC_CCR_01012026_v2.xlsx'
    Sanitiza el sufijo (sin espacios ni caracteres problemáticos para filename).
    """
    if not sufijo:
        return nombre
    # Sanitización: solo letras, dígitos, guiones bajos y guiones
    limpio = "".join(c for c in str(sufijo).strip() if c.isalnum() or c in ("_", "-"))
    if not limpio:
        return nombre
    if "." in nombre:
        base, ext = nombre.rsplit(".", 1)
        return f"{base}_{limpio}.{ext}"
    return f"{nombre}_{limpio}"


# =========================
# HEADER
# =========================
st.markdown("""
<div class="header-block">
  <h1>🛢 Generador y Validador -- Ordenes Corpored </h1>
  <p>Validador · Generador · Captura manual — CCR / FRAVAL</p>
</div>
""", unsafe_allow_html=True)


# =========================
# AUTENTICACIÓN
# =========================
# Credenciales: en producción se leen de st.secrets (panel de Streamlit Cloud).
# Para corrida local, si no hay secrets configurados, usar este default.
# Formato esperado en secrets.toml:
#     [credenciales]
#     didier  = "claveDidier123"
#     juan    = "claveJuan456"
DEFAULT_CREDENCIALES = {
    "didier": "cambiar123",
    # agrega aquí los compañeros que necesiten acceso en modo local
}

def _cargar_credenciales():
    """Lee credenciales de st.secrets si existen, sino devuelve el default."""
    try:
        if "credenciales" in st.secrets:
            return {str(k): str(v) for k, v in st.secrets["credenciales"].items()}
    except (FileNotFoundError, Exception):
        pass
    return DEFAULT_CREDENCIALES

def _pedir_login():
    """Muestra el formulario de login y bloquea el resto de la app hasta autenticar."""
    st.markdown('<div class="seccion-titulo">▸ Acceso</div>', unsafe_allow_html=True)
    st.caption("Ingresa tu usuario y contraseña para usar la aplicación.")

    with st.form("login_form", clear_on_submit=False):
        usuario = st.text_input("Usuario", key="login_usr").strip().lower()
        clave   = st.text_input("Contraseña", type="password", key="login_pwd")
        submit  = st.form_submit_button("🔓 Entrar", use_container_width=True)

    if submit:
        creds = _cargar_credenciales()
        if usuario in creds and creds[usuario] == clave:
            st.session_state["auth_user"] = usuario
            st.rerun()
        else:
            st.error("❌ Usuario o contraseña incorrectos.")
    st.stop()

# Si no está autenticado, mostrar login y detener todo lo demás
if "auth_user" not in st.session_state:
    _pedir_login()

# Barra superior con usuario logueado + cerrar sesión
_col_user, _col_logout = st.columns([5, 1])
with _col_user:
    st.caption(f"👤 Sesión: **{st.session_state['auth_user']}**")
with _col_logout:
    if st.button("Cerrar sesión", key="btn_logout", use_container_width=True):
        for k in ("auth_user", "login_usr", "login_pwd"):
            st.session_state.pop(k, None)
        st.rerun()


# =========================
# UI PRINCIPAL
# =========================
modo  = st.selectbox("Modo:", ["Selecciona una opción", "Validar", "Generar"], key="modo")
fecha = st.date_input("📅 Selecciona la fecha", value=date.today(), key="fecha")

if modo == "Selecciona una opción":
    st.info("👆 Selecciona un modo para continuar")
    st.stop()

fecha_usuario = fecha.strftime("%d%m%Y")
archivo = st.file_uploader("📂 Sube tu archivo Excel", type=["xlsx", "xlsm"], key="archivo")

# Invalidar caché si cambia archivo o fecha
if archivo is not None:
    archivo_id = f"{archivo.name}_{archivo.size}_{fecha_usuario}"
    if st.session_state.get("archivo_id") != archivo_id:
        for k in ["ccr_file", "df_auto", "df_firme_extra", "archivo_id", "archivo_error"]:
            st.session_state.pop(k, None)
        st.session_state["archivo_id"] = archivo_id

    if "df_auto" not in st.session_state and "archivo_error" not in st.session_state:
        try:
            df_a, df_e = procesar(archivo)
            st.session_state["df_auto"] = df_a
            st.session_state["df_firme_extra"] = df_e
        except ExcelStructureError as e:
            st.session_state["archivo_error"] = str(e)
            st.session_state["df_auto"] = None
            st.session_state["df_firme_extra"] = None
            logger.error("ExcelStructureError: %s", e)
        except Exception as e:
            st.session_state["archivo_error"] = (
                f"Error inesperado leyendo el archivo: {e}"
            )
            st.session_state["df_auto"] = None
            st.session_state["df_firme_extra"] = None
            logger.error("Error inesperado en procesar(): %s\n%s", e, traceback.format_exc())

# Mostrar el error de archivo (si existe) en cualquier modo
if archivo is not None and st.session_state.get("archivo_error"):
    st.error(f"❌ {st.session_state['archivo_error']}")


# ══════════════════════════════════════════════════════
# MODO VALIDAR
# ══════════════════════════════════════════════════════
if modo == "Validar":

    if archivo is None:
        st.info("📂 Sube el archivo Excel para continuar")
        st.stop()

    if st.session_state.get("archivo_error"):
        st.stop()  # ya se mostró arriba

    tab_val_faltantes, tab_val_extra = st.tabs([
        "📊 Faltantes",
        "📤 Ordenes Extra"
    ])

    # ── TAB FALTANTES (flujo de siempre) ────────────────
    with tab_val_faltantes:
        df_auto = st.session_state.get("df_auto")

        if df_auto is None:
            st.success("✅ Programa completo, sin faltantes")
        else:
            df_ccr = df_auto[df_auto["#"] < LIMITE_FRAVAL]

            if df_ccr.empty:
                st.success("✅ Ordenes Completas")
            else:
                st.subheader("📊 Destinos faltantes (CCR)")
                df_mostrar = df_ccr[["#", "Nombre Prod", "Tipo", "Numero de Tonel"]].copy()
                df_mostrar.columns = ["# Destino", "Producto", "Tipo", "# Tonel"]
                st.dataframe(df_mostrar, use_container_width=True)

                opcion = st.radio("¿Deseas generar los archivos?", ["No", "Sí"])
                if opcion == "Sí":
                    if "ccr_file" not in st.session_state:
                        try:
                            salida = construir(df_ccr, fecha_usuario, CLIENTE_CCR)
                            st.session_state["ccr_file"] = generar_excel(salida)
                            st.success("✅ Archivo listo para descargar")
                        except Exception as e:
                            logger.error("Error generando Excel CCR: %s\n%s", e, traceback.format_exc())
                            st.error(f"❌ Error generando el archivo: {e}")

                if "ccr_file" in st.session_state:
                    sufijo_val = st.text_input(
                        "Sufijo (opcional)",
                        placeholder="ej: v2, v3, REVISADO — déjalo vacío si no aplica",
                        key="sufijo_val",
                        help="Se agregará al final del nombre del archivo, antes de la extensión."
                    )
                    nombre_descarga = aplicar_sufijo(f"SIIC_CCR_{fecha_usuario}.xlsx", sufijo_val)
                    st.download_button(
                        "⬇️ Descargar CCR",
                        st.session_state["ccr_file"],
                        nombre_descarga,
                        mime=MIME_XLSX
                    )

    # ── TAB FIRME EXTRA (Validar) ───────────────────────
    with tab_val_extra:
        st.markdown('<div class="seccion-titulo">▸ Órdenes en FIRME que no están en Programa</div>',
                    unsafe_allow_html=True)
        st.caption("Filas de FIRME que cumplen los requisitos de litros (≥32,000 TN, "
                   "exactos para compartimentos) pero NO tienen una orden equivalente en "
                   "Programa del día ni en Compartimentos.")

        df_extra = st.session_state.get("df_firme_extra")
        if df_extra is None or df_extra.empty:
            st.success("✅ Todo en FIRME tiene su contraparte en Programa/Compartimentos.")
        else:
            n_ccr_ex    = len(df_extra[df_extra["#"] < LIMITE_FRAVAL])
            n_fraval_ex = len(df_extra[df_extra["#"] >= LIMITE_FRAVAL])
            col1, col2, col3 = st.columns(3)
            col1.metric("Total Ordenes Extra",                  len(df_extra))
            col2.metric(f"CCR (dest < {LIMITE_FRAVAL})",      n_ccr_ex)
            col3.metric(f"FRAVAL (dest ≥ {LIMITE_FRAVAL})",   n_fraval_ex)
            st.markdown("---")

            # Resumen por destinos fijos
            resumen_html = render_resumen_firme_extra_html(df_extra)
            if resumen_html:
                st.markdown(
                    f'<div class="seccion-titulo">▸ Resumen por destino '
                    f'({", ".join(str(d) for d in DESTINOS_RESUMEN_FIRME_EXTRA)})</div>',
                    unsafe_allow_html=True
                )
                st.markdown(resumen_html, unsafe_allow_html=True)
                st.markdown("&nbsp;")

            # Tabla detallada (omite las filas que ya están en el resumen:
            # destinos fijos + tonel 1 o 2). Las de compartimentos (tonel 3 o 4)
            # a esos destinos sí aparecen aquí.
            df_detalle = df_extra[~(
                df_extra["#"].isin(DESTINOS_RESUMEN_FIRME_EXTRA) &
                df_extra["Numero de Tonel"].astype(int).isin([1, 2])
            )]

            st.markdown('<div class="seccion-titulo">▸ Detalle completo</div>',
                        unsafe_allow_html=True)
            if df_detalle.empty:
                st.caption("Todas las órdenes adicionales están reflejadas en el resumen.")
            else:
                filas_html = ""
                for _, r in df_detalle.iterrows():
                    p = r["Nombre Prod"]; bg = color_prod(p); fg = color_text(p)
                    badge = (f'<span class="badge-ccr">CCR</span>'
                             if r["#"] < LIMITE_FRAVAL
                             else f'<span class="badge-fraval">FRAVAL</span>')
                    filas_html += f"""<tr>
                        <td>{r['#']}</td>
                        <td style="background:{bg};color:{fg};">{p}</td>
                        <td>{r['Numero de Tonel']}</td>
                        <td>{r['Tipo']}</td>
                        <td>{badge}</td></tr>"""
                st.markdown(f"""<table class="preview-table">
                    <tr><th>#</th><th>Producto</th><th>Tonel</th><th>Tipo</th><th>Cliente</th></tr>
                    {filas_html}</table>""", unsafe_allow_html=True)

            st.markdown("---")
            st.markdown('<div class="seccion-titulo">▸ Exportar Excel (Ordenes Extra)</div>',
                        unsafe_allow_html=True)

            sufijo_extra_val = st.text_input(
                "Sufijo (opcional)",
                placeholder="ej: v2, EXTRA — déjalo vacío si no aplica",
                key="sufijo_extra_val",
                help="Se agregará al final del nombre del archivo."
            )

            try:
                nombre_extra_val = aplicar_sufijo(
                    f"ORDENES_EXTRA_{fecha_usuario}.xlsx",
                    sufijo_extra_val
                )
                bytes_extra_val = generar_excel_firme_extra(df_extra, fecha_usuario)
                st.download_button(
                    f"📥 Descargar {nombre_extra_val}",
                    data=bytes_extra_val,
                    file_name=nombre_extra_val,
                    mime=MIME_XLSX,
                    use_container_width=True,
                    key="dl_extra_val"
                )
            except Exception as e:
                logger.error("Error generando Excel Ordenes Extra (Validar): %s\n%s",
                             e, traceback.format_exc())
                st.error(f"❌ Error generando el archivo Excel: {e}")


# ══════════════════════════════════════════════════════
# MODO GENERAR — tabs: Automático + captura manual + preview
# ══════════════════════════════════════════════════════
elif modo == "Generar":

    tab_auto, tab_tn, tab_comp, tab_extra, tab_preview = st.tabs([
        "📁 Automático (Excel)",
        "🚛 Tonel Normal",
        "🔧 Compartimentos",
        "📤 Ordenes Extra",
        "📋 Vista Previa & Exportar"
    ])

    # ── TAB AUTOMÁTICO ──────────────────────────────────
    with tab_auto:
        st.markdown('<div class="seccion-titulo">▸ Faltantes detectados del archivo Excel</div>',
                    unsafe_allow_html=True)

        if archivo is None:
            st.info("📂 Sube el archivo Excel arriba para detectar faltantes automáticamente.")
            st.caption("También puedes ir a las otras pestañas para capturar registros manualmente.")
        elif st.session_state.get("archivo_error"):
            st.warning("Corrige el error del archivo arriba para ver los faltantes.")
        else:
            df_auto = st.session_state.get("df_auto")
            if df_auto is None:
                st.success("✅ No se detectaron faltantes en el archivo.")
            else:
                n_ccr    = len(df_auto[df_auto["#"] < LIMITE_FRAVAL])
                n_fraval = len(df_auto[df_auto["#"] >= LIMITE_FRAVAL])
                col1, col2, col3 = st.columns(3)
                col1.metric("Total faltantes",               len(df_auto))
                col2.metric(f"CCR (dest < {LIMITE_FRAVAL})",  n_ccr)
                col3.metric(f"FRAVAL (dest ≥ {LIMITE_FRAVAL})", n_fraval)
                st.markdown("---")

                filas_html = ""
                for _, r in df_auto.iterrows():
                    p = r["Nombre Prod"]; bg = color_prod(p); fg = color_text(p)
                    badge = (f'<span class="badge-ccr">CCR</span>'
                             if r["#"] < LIMITE_FRAVAL
                             else f'<span class="badge-fraval">FRAVAL</span>')
                    filas_html += f"""<tr>
                        <td>{r['#']}</td>
                        <td style="background:{bg};color:{fg};">{p}</td>
                        <td>{r['Numero de Tonel']}</td>
                        <td>{r['Tipo']}</td>
                        <td>{badge}</td></tr>"""
                st.markdown(f"""<table class="preview-table">
                    <tr><th>#</th><th>Producto</th><th>Tonel</th><th>Tipo</th><th>Cliente</th></tr>
                    {filas_html}</table>""", unsafe_allow_html=True)

    # ── TAB TONEL NORMAL ────────────────────────────────
    with tab_tn:

        # Selectores de unidad y operador (afectan Individual + Fulles)
        st.markdown('<div class="seccion-titulo">▸ Unidad y Operador</div>', unsafe_allow_html=True)
        st.caption("Aplica a todas las órdenes de Individual y Fulles. Cambiar reescribe las ya capturadas.")

        sel_col1, sel_col2 = st.columns(2)
        with sel_col1:
            vehiculo_tn_sel = st.selectbox(
                "Unidad", VEHICULOS_TN_OPCIONES,
                index=VEHICULOS_TN_OPCIONES.index(st.session_state["vehiculo_tn"])
                       if st.session_state["vehiculo_tn"] in VEHICULOS_TN_OPCIONES else 0,
                key="vehiculo_tn_sel"
            )
        with sel_col2:
            rfc_tn_sel = st.selectbox(
                "Operador (RFC)", RFC_CHOFER_OPCIONES,
                index=RFC_CHOFER_OPCIONES.index(st.session_state["rfc_tn"])
                       if st.session_state["rfc_tn"] in RFC_CHOFER_OPCIONES else 0,
                key="rfc_tn_sel"
            )

        # Si cambió, reescribir las órdenes ya capturadas y persistir en estado
        cambios = []
        if vehiculo_tn_sel != st.session_state["vehiculo_tn"]:
            st.session_state["vehiculo_tn"] = vehiculo_tn_sel
            for reg in st.session_state.tn_individual:
                reg["Clave de Vehiculo"] = vehiculo_tn_sel
            for reg in st.session_state.tn_masivo:
                reg["Clave de Vehiculo"] = vehiculo_tn_sel
            cambios.append(f"Unidad → {vehiculo_tn_sel}")
        if rfc_tn_sel != st.session_state["rfc_tn"]:
            st.session_state["rfc_tn"] = rfc_tn_sel
            for reg in st.session_state.tn_individual:
                reg["R.F.C. chofer"] = rfc_tn_sel
            for reg in st.session_state.tn_masivo:
                reg["R.F.C. chofer"] = rfc_tn_sel
            cambios.append(f"Operador → {rfc_tn_sel}")
        if cambios:
            n_total = len(st.session_state.tn_individual) + len(st.session_state.tn_masivo)
            if n_total > 0:
                st.info(f"🔁 Reescribiendo {n_total} orden(es) capturada(s): " + " · ".join(cambios))

        st.markdown("---")

        # Individual
        st.markdown('<div class="seccion-titulo">▸ Individual</div>', unsafe_allow_html=True)
        st.caption("Agrega una orden a la vez.")

        col1, col2, col3 = st.columns(3)
        with col1:
            st.markdown('<div class="no-stepper-anchor"></div>', unsafe_allow_html=True)
            tn_destino = st.number_input(
                "Destino", min_value=1, max_value=DESTINO_MAX,
                value=None, step=1, format="%d",
                placeholder="Ej: 26", key="tn_dest"
            )
        with col2:
            tn_producto = st.selectbox("Producto", OPCIONES_PROD, key="tn_prod")
        with col3:
            tn_tonel = st.selectbox("Tonel", sorted(TONELES_TN_VALIDOS), key="tn_ton")

        if st.button("➕ Agregar Orden", key="btn_tn_ind"):
            try:
                d_int = validar_destino(tn_destino, etiqueta="Destino")
                p_can = validar_producto(tn_producto, etiqueta="Producto")
                t_int = validar_tonel_tn(tn_tonel, etiqueta="Tonel")
            except ValidationError as ve:
                st.error(str(ve))
            else:
                st.session_state.tn_individual.append({
                    "#": d_int, "Nombre Prod": p_can,
                    "Clve de producto": MAPA_PRODUCTO[p_can],
                    "Numero de Tonel": t_int,
                    "Clave de Vehiculo": st.session_state["vehiculo_tn"],
                    "R.F.C. chofer":     st.session_state["rfc_tn"],
                    "Tipo": "NORMAL", "origen": "manual"
                })
                st.success(f"✅ Registro agregado (Destino {d_int} · {p_can} · Tonel {t_int})")
                st.rerun()

        if st.session_state.tn_individual:
            # Encabezado tipo tabla
            st.markdown(
                '<div style="background:#1a1e2e;color:#e8a020;'
                'padding:8px 12px;border-radius:6px 6px 0 0;'
                'font-size:0.7rem;letter-spacing:1.5px;text-transform:uppercase;'
                'font-weight:700;display:flex;align-items:center;">'
                '<div style="flex:1;text-align:center;">#</div>'
                '<div style="flex:2;text-align:center;">Producto</div>'
                '<div style="flex:1.5;text-align:center;">Clave</div>'
                '<div style="flex:1;text-align:center;">Tonel</div>'
                '<div style="flex:2;text-align:center;">Vehículo</div>'
                '<div style="flex:1;text-align:center;">✏️</div>'
                '<div style="flex:1;text-align:center;">🗑️</div>'
                '</div>',
                unsafe_allow_html=True
            )

            # Filas interactivas
            for idx, registro in enumerate(st.session_state.tn_individual):
                p  = registro["Nombre Prod"]
                bg = color_prod(p)
                fg = color_text(p)
                col_n, col_p, col_c, col_t, col_v, col_e, col_d = st.columns([1, 2, 1.5, 1, 2, 1, 1])
                col_n.markdown(
                    f"<div style='text-align:center;padding:6px 0;font-weight:500;'>"
                    f"{registro['#']}</div>",
                    unsafe_allow_html=True
                )
                col_p.markdown(
                    f"<div style='text-align:center;padding:6px 0;background:{bg};color:{fg};"
                    f"border-radius:4px;font-weight:600;'>{p}</div>",
                    unsafe_allow_html=True
                )
                col_c.markdown(
                    f"<div style='text-align:center;padding:6px 0;font-weight:500;'>"
                    f"{registro['Clve de producto']}</div>",
                    unsafe_allow_html=True
                )
                col_t.markdown(
                    f"<div style='text-align:center;padding:6px 0;font-weight:500;'>"
                    f"{registro['Numero de Tonel']}</div>",
                    unsafe_allow_html=True
                )
                col_v.markdown(
                    f"<div style='text-align:center;padding:6px 0;font-weight:500;'>"
                    f"{registro['Clave de Vehiculo']}</div>",
                    unsafe_allow_html=True
                )
                if col_e.button("✏️", key=f"edit_ind_{idx}", help="Editar esta orden"):
                    st.session_state["editando_ind"] = idx
                    st.rerun()
                if col_d.button("🗑️", key=f"del_ind_{idx}", help="Eliminar esta orden"):
                    st.session_state.tn_individual.pop(idx)
                    # Si estábamos editando esta misma fila u otra posterior, recalibrar
                    if st.session_state.get("editando_ind") == idx:
                        st.session_state.pop("editando_ind", None)
                    elif st.session_state.get("editando_ind", -1) > idx:
                        st.session_state["editando_ind"] -= 1
                    st.rerun()

            # Formulario de edición inline
            edit_idx = st.session_state.get("editando_ind")
            if edit_idx is not None and 0 <= edit_idx < len(st.session_state.tn_individual):
                st.markdown("---")
                st.markdown(
                    f'<div class="seccion-titulo">▸ Editando orden #{edit_idx + 1}</div>',
                    unsafe_allow_html=True
                )
                actual = st.session_state.tn_individual[edit_idx]

                ec1, ec2, ec3 = st.columns(3)
                with ec1:
                    st.markdown('<div class="no-stepper-anchor"></div>', unsafe_allow_html=True)
                    e_destino = st.number_input(
                        "Destino", min_value=1, max_value=DESTINO_MAX,
                        value=int(actual["#"]), step=1, format="%d",
                        key=f"edit_dest_{edit_idx}"
                    )
                with ec2:
                    prod_idx = OPCIONES_PROD.index(actual["Nombre Prod"]) \
                               if actual["Nombre Prod"] in OPCIONES_PROD else 0
                    e_producto = st.selectbox(
                        "Producto", OPCIONES_PROD, index=prod_idx,
                        key=f"edit_prod_{edit_idx}"
                    )
                with ec3:
                    toneles_lista = sorted(TONELES_TN_VALIDOS)
                    ton_idx = toneles_lista.index(int(actual["Numero de Tonel"])) \
                              if int(actual["Numero de Tonel"]) in toneles_lista else 0
                    e_tonel = st.selectbox(
                        "Tonel", toneles_lista, index=ton_idx,
                        key=f"edit_ton_{edit_idx}"
                    )

                bg1, bg2, _ = st.columns([2, 2, 6])
                with bg1:
                    if st.button("💾 Guardar cambios", key=f"save_ind_{edit_idx}",
                                 use_container_width=True):
                        try:
                            d_int = validar_destino(e_destino, etiqueta="Destino")
                            p_can = validar_producto(e_producto, etiqueta="Producto")
                            t_int = validar_tonel_tn(e_tonel, etiqueta="Tonel")
                        except ValidationError as ve:
                            st.error(str(ve))
                        else:
                            st.session_state.tn_individual[edit_idx] = {
                                "#": d_int, "Nombre Prod": p_can,
                                "Clve de producto": MAPA_PRODUCTO[p_can],
                                "Numero de Tonel": t_int,
                                "Clave de Vehiculo": st.session_state["vehiculo_tn"],
                                "R.F.C. chofer":     st.session_state["rfc_tn"],
                                "Tipo": "NORMAL", "origen": "manual"
                            }
                            st.session_state.pop("editando_ind", None)
                            st.success("✅ Orden actualizada")
                            st.rerun()
                with bg2:
                    if st.button("↩️ Cancelar", key=f"cancel_ind_{edit_idx}",
                                 use_container_width=True):
                        st.session_state.pop("editando_ind", None)
                        st.rerun()

            # Botón limpiar todo (se conserva)
            st.markdown("&nbsp;")
            if st.button("🗑 Limpiar individuales", key="limpiar_ind"):
                st.session_state.tn_individual = []
                st.session_state.pop("editando_ind", None)
                st.rerun()

        st.markdown("---")

        # Masivo
        st.markdown('<div class="seccion-titulo">▸ Fulles</div>', unsafe_allow_html=True)
        st.caption("Genera N fulles para el mismo destino y producto. Cada full = tonel 1 + tonel 2.")

        col1, col2, col3 = st.columns(3)
        with col1: mas_destino  = st.text_input("Destino",   placeholder="ej: 87", key="mas_dest")
        with col2: mas_producto = st.selectbox("Producto",   OPCIONES_PROD,        key="mas_prod")
        with col3: mas_cantidad = st.number_input("# Fulles", min_value=1, max_value=99, value=1, key="mas_cant")

        if st.button("➕ Agregar fulles", key="btn_masivo"):
            try:
                dest_int = validar_destino(mas_destino, etiqueta="Destino")
                # Producto ya viene del selectbox, pero validamos por consistencia
                prod_can = validar_producto(mas_producto, etiqueta="Producto")
            except ValidationError as ve:
                st.error(str(ve))
            else:
                nuevos = []
                for _ in range(int(mas_cantidad)):
                    for tonel in [1, 2]:
                        nuevos.append({
                            "#": dest_int, "Nombre Prod": prod_can,
                            "Clve de producto": MAPA_PRODUCTO[prod_can],
                            "Numero de Tonel": tonel,
                            "Clave de Vehiculo": st.session_state["vehiculo_tn"],
                            "R.F.C. chofer":     st.session_state["rfc_tn"],
                            "Tipo": "NORMAL", "origen": "manual"
                        })
                st.session_state.tn_masivo.extend(nuevos)
                st.success(f"✅ {len(nuevos)} órdenes agregadas ({int(mas_cantidad)} full(es) × 2 toneles)")
                st.rerun()

        if st.session_state.tn_masivo:
            df_mas = pd.DataFrame(st.session_state.tn_masivo)
            filas_html = ""
            for _, r in df_mas.iterrows():
                p = r["Nombre Prod"]; bg = color_prod(p); fg = color_text(p)
                filas_html += f"""<tr>
                    <td>{r['#']}</td>
                    <td style="background:{bg};color:{fg};">{p}</td>
                    <td>{r['Numero de Tonel']}</td></tr>"""
            st.markdown(f"""<table class="tabla-camion">
                <tr><th>#</th><th>Producto</th><th>Tonel</th></tr>
                {filas_html}</table>""", unsafe_allow_html=True)
            if st.button("🗑 Limpiar masivo", key="limpiar_mas"):
                st.session_state.tn_masivo = []
                st.rerun()

    # ── TAB COMPARTIMENTOS ──────────────────────────────
    with tab_comp:
        # Selectores de unidad y operador para compartimentos
        st.markdown('<div class="seccion-titulo">▸ Unidad y Operador</div>', unsafe_allow_html=True)
        st.caption("Aplica a todos los compartimentos. Cambiar reescribe las ya capturadas.")

        sel_col1, sel_col2 = st.columns(2)
        with sel_col1:
            vehiculo_tc_sel = st.selectbox(
                "Unidad", VEHICULOS_TC_OPCIONES,
                index=VEHICULOS_TC_OPCIONES.index(st.session_state["vehiculo_tc"])
                       if st.session_state["vehiculo_tc"] in VEHICULOS_TC_OPCIONES else 0,
                key="vehiculo_tc_sel"
            )
        with sel_col2:
            rfc_tc_sel = st.selectbox(
                "Operador (RFC)", RFC_CHOFER_OPCIONES,
                index=RFC_CHOFER_OPCIONES.index(st.session_state["rfc_tc"])
                       if st.session_state["rfc_tc"] in RFC_CHOFER_OPCIONES else 0,
                key="rfc_tc_sel"
            )

        # Si cambió, persistir y notificar (la reescritura se hace al recolectar/exportar)
        cambios_tc = []
        if vehiculo_tc_sel != st.session_state["vehiculo_tc"]:
            st.session_state["vehiculo_tc"] = vehiculo_tc_sel
            cambios_tc.append(f"Unidad → {vehiculo_tc_sel}")
        if rfc_tc_sel != st.session_state["rfc_tc"]:
            st.session_state["rfc_tc"] = rfc_tc_sel
            cambios_tc.append(f"Operador → {rfc_tc_sel}")
        if cambios_tc and st.session_state.compartimentos_tc:
            n = len(st.session_state.compartimentos_tc) * 4
            st.info(f"🔁 Aplicado a {n} orden(es) de compartimentos: " + " · ".join(cambios_tc))

        st.markdown("---")

        st.markdown('<div class="seccion-titulo">▸ Agregar Compartimentos</div>', unsafe_allow_html=True)
        st.caption("Selecciona N/A en un compartimento si no quieres generar esa orden. "
                   "Si todos los compartimentos de un tonel son N/A, ese tonel no se genera.")

        col_t1, col_sep, col_t2 = st.columns([5, 1, 5])
        _idx_na = OPCIONES_PROD_COMP.index(NA_VALUE)
        with col_t1:
            st.markdown("**🔵 Tonel 1** — Destino compartimentos 1 y 2")
            tc_dest_t1 = st.text_input("Destino T1", placeholder="ej: 87 (opcional si C1 y C2 son N/A)",  key="tc_d1")
            c1, c2 = st.columns(2)
            with c1: tc_prod_c1 = st.selectbox("Comp 1", OPCIONES_PROD_COMP, index=_idx_na, key="tc_c1")
            with c2: tc_prod_c2 = st.selectbox("Comp 2", OPCIONES_PROD_COMP, index=_idx_na, key="tc_c2")
        with col_sep:
            st.markdown("<div style='text-align:center;color:#aaa;padding-top:60px;font-size:1.5rem;'>│</div>",
                        unsafe_allow_html=True)
        with col_t2:
            st.markdown("**🟠 Tonel 2** — Destino compartimentos 3 y 4")
            tc_dest_t2 = st.text_input("Destino T2", placeholder="ej: 349 (opcional si C3 y C4 son N/A)", key="tc_d2")
            c3, c4 = st.columns(2)
            with c3: tc_prod_c3 = st.selectbox("Comp 3", OPCIONES_PROD_COMP, index=_idx_na, key="tc_c3")
            with c4: tc_prod_c4 = st.selectbox("Comp 4", OPCIONES_PROD_COMP, index=_idx_na, key="tc_c4")

        col_btn1, col_btn2, _ = st.columns([2, 2, 6])
        with col_btn1: agregar_tc   = st.button("➕ Agregar compartimentos", key="btn_agregar_tc", use_container_width=True)
        with col_btn2: limpiar_tc_b = st.button("🗑 Limpiar lista",  key="btn_limpiar_tc", use_container_width=True)

        if agregar_tc:
            errores = []
            # ¿Algún compartimento de cada tonel tiene producto real (≠ N/A)?
            t1_activo = tc_prod_c1 != NA_VALUE or tc_prod_c2 != NA_VALUE
            t2_activo = tc_prod_c3 != NA_VALUE or tc_prod_c4 != NA_VALUE

            # ¿El usuario llenó cada destino?
            t1_dest_lleno = bool(str(tc_dest_t1).strip())
            t2_dest_lleno = bool(str(tc_dest_t2).strip())

            # Si los 4 son N/A → bloquear (vale aunque destinos estén vacíos)
            if not t1_activo and not t2_activo:
                errores.append(
                    "No puede haber un compartimento totalmente vacío "
                    "(los 4 productos no pueden ser N/A)."
                )

            # Tonel 1: validar combinación destino vs productos
            d1 = None
            if t1_activo and not t1_dest_lleno:
                errores.append(
                    "Destino T1 vacío pero hay productos seleccionados en Comp 1 o Comp 2."
                )
            elif t1_activo and t1_dest_lleno:
                try:
                    d1 = validar_destino(tc_dest_t1, etiqueta="Destino T1")
                except ValidationError as ve:
                    errores.append(str(ve))
            elif not t1_activo and t1_dest_lleno:
                # Destino con valor pero ambos comps son N/A → forzar N/A en el tonel
                # (silenciosamente: el destino se ignora)
                pass

            # Tonel 2: misma lógica
            d2 = None
            if t2_activo and not t2_dest_lleno:
                errores.append(
                    "Destino T2 vacío pero hay productos seleccionados en Comp 3 o Comp 4."
                )
            elif t2_activo and t2_dest_lleno:
                try:
                    d2 = validar_destino(tc_dest_t2, etiqueta="Destino T2")
                except ValidationError as ve:
                    errores.append(str(ve))
            elif not t2_activo and t2_dest_lleno:
                pass

            if errores:
                for e in errores:
                    st.error(e)
            else:
                st.session_state.compartimentos_tc.append({
                    "d1": d1, "d2": d2,
                    "prods": [tc_prod_c1, tc_prod_c2, tc_prod_c3, tc_prod_c4]
                })
                st.success(f"✅ Compartimento #{len(st.session_state.compartimentos_tc)} agregado")
                st.rerun()

        if limpiar_tc_b:
            st.session_state.compartimentos_tc = []
            st.rerun()

        if st.session_state.compartimentos_tc:
            st.markdown("---")
            st.markdown(f'<div class="seccion-titulo">▸ Compartimentos cargados ({len(st.session_state.compartimentos_tc)})</div>',
                        unsafe_allow_html=True)
            filas = ""
            total_ordenes = 0
            for i, cam in enumerate(st.session_state.compartimentos_tc):
                p = cam["prods"]
                # Contar órdenes reales (compartimentos con producto ≠ N/A)
                total_ordenes += sum(1 for x in p if x != NA_VALUE)
                def celda(prod, num):
                    if prod == NA_VALUE:
                        return (
                            f'<td style="background:#F4F5F7;color:#9CA3AF;padding:5px 10px;'
                            f'border:1px solid #eee;text-align:center;font-style:italic;">'
                            f'C{num}: N/A</td>'
                        )
                    bg = color_prod(prod); fg = color_text(prod)
                    return f'<td style="background:{bg};color:{fg};padding:5px 10px;border:1px solid #eee;text-align:center;">C{num}: {prod}</td>'

                # Si el destino es None (todo el tonel es N/A), mostrar "—"
                d1_show = cam['d1'] if cam['d1'] is not None else "—"
                d2_show = cam['d2'] if cam['d2'] is not None else "—"
                filas += f"""<tr>
                    <td style="padding:5px 10px;border:1px solid #eee;text-align:center;color:#e8a020;font-weight:700;">{i+1}</td>
                    <td style="padding:5px 10px;border:1px solid #eee;text-align:center;">{d1_show}</td>
                    {celda(p[0],1)}{celda(p[1],2)}
                    <td style="padding:5px 10px;border:1px solid #eee;text-align:center;">{d2_show}</td>
                    {celda(p[2],3)}{celda(p[3],4)}
                </tr>"""
            st.markdown(f"""<table class="tabla-camion" style="font-size:0.8rem;">
                <tr><th>#</th><th>Dest T1</th><th>Comp 1</th><th>Comp 2</th><th>Dest T2</th><th>Comp 3</th><th>Comp 4</th></tr>
                {filas}</table>
                <small style="color:#666;display:block;margin-top:6px;">
                    {len(st.session_state.compartimentos_tc)} compartimento(s) · {total_ordenes} orden(es)
                </small>""", unsafe_allow_html=True)

            st.markdown("&nbsp;")
            idx_del = st.number_input("Eliminar compartimento #", min_value=1,
                                       max_value=len(st.session_state.compartimentos_tc),
                                       value=1, key="idx_del")
            if st.button("❌ Eliminar ese compartimento", key="btn_del_cam"):
                st.session_state.compartimentos_tc.pop(int(idx_del) - 1)
                st.success("Compartimento eliminado")
                st.rerun()

    # ── TAB FIRME EXTRA ─────────────────────────────────
    with tab_extra:
        st.markdown('<div class="seccion-titulo">▸ Órdenes en FIRME que no están en Programa</div>',
                    unsafe_allow_html=True)
        st.caption("Filas de FIRME que cumplen los requisitos de litros (≥32,000 TN, "
                   "exactos para compartimentos) pero NO tienen una orden equivalente en "
                   "Programa del día ni en Compartimentos.")

        if archivo is None:
            st.info("📂 Sube el archivo Excel arriba para ver las órdenes extra de FIRME.")
        elif st.session_state.get("archivo_error"):
            st.warning("Corrige el error del archivo arriba para ver las órdenes extra.")
        else:
            df_extra = st.session_state.get("df_firme_extra")
            if df_extra is None or df_extra.empty:
                st.success("✅ Todo en FIRME tiene su contraparte en Programa/Compartimentos.")
            else:
                n_ccr_ex    = len(df_extra[df_extra["#"] < LIMITE_FRAVAL])
                n_fraval_ex = len(df_extra[df_extra["#"] >= LIMITE_FRAVAL])
                col1, col2, col3 = st.columns(3)
                col1.metric("Total Ordenes Extra",                  len(df_extra))
                col2.metric(f"CCR (dest < {LIMITE_FRAVAL})",      n_ccr_ex)
                col3.metric(f"FRAVAL (dest ≥ {LIMITE_FRAVAL})",   n_fraval_ex)
                st.markdown("---")

                # Resumen por destinos fijos
                resumen_html = render_resumen_firme_extra_html(df_extra)
                if resumen_html:
                    st.markdown(
                        f'<div class="seccion-titulo">▸ Resumen por destino '
                        f'({", ".join(str(d) for d in DESTINOS_RESUMEN_FIRME_EXTRA)})</div>',
                        unsafe_allow_html=True
                    )
                    st.markdown(resumen_html, unsafe_allow_html=True)
                    st.markdown("&nbsp;")

                # Tabla detallada (omite las filas que ya están en el resumen:
                # destinos fijos + tonel 1 o 2). Las de compartimentos (tonel 3 o 4)
                # a esos destinos sí aparecen aquí.
                df_detalle = df_extra[~(
                    df_extra["#"].isin(DESTINOS_RESUMEN_FIRME_EXTRA) &
                    df_extra["Numero de Tonel"].astype(int).isin([1, 2])
                )]

                st.markdown('<div class="seccion-titulo">▸ Detalle completo</div>',
                            unsafe_allow_html=True)
                if df_detalle.empty:
                    st.caption("Todas las órdenes adicionales están reflejadas en el resumen.")
                else:
                    filas_html = ""
                    for _, r in df_detalle.iterrows():
                        p = r["Nombre Prod"]; bg = color_prod(p); fg = color_text(p)
                        badge = (f'<span class="badge-ccr">CCR</span>'
                                 if r["#"] < LIMITE_FRAVAL
                                 else f'<span class="badge-fraval">FRAVAL</span>')
                        filas_html += f"""<tr>
                            <td>{r['#']}</td>
                            <td style="background:{bg};color:{fg};">{p}</td>
                            <td>{r['Numero de Tonel']}</td>
                            <td>{r['Tipo']}</td>
                            <td>{badge}</td></tr>"""
                    st.markdown(f"""<table class="preview-table">
                        <tr><th>#</th><th>Producto</th><th>Tonel</th><th>Tipo</th><th>Cliente</th></tr>
                        {filas_html}</table>""", unsafe_allow_html=True)

                st.markdown("---")
                st.markdown('<div class="seccion-titulo">▸ Exportar Excel (Ordenes Extra)</div>',
                            unsafe_allow_html=True)

                sufijo_extra = st.text_input(
                    "Sufijo (opcional)",
                    placeholder="ej: v2, EXTRA — déjalo vacío si no aplica",
                    key="sufijo_extra",
                    help="Se agregará al final del nombre del archivo."
                )

                try:
                    nombre_extra = aplicar_sufijo(
                        f"ORDENES_EXTRA_{fecha_usuario}.xlsx",
                        sufijo_extra
                    )
                    bytes_extra = generar_excel_firme_extra(df_extra, fecha_usuario)
                    st.download_button(
                        f"📥 Descargar {nombre_extra}",
                        data=bytes_extra,
                        file_name=nombre_extra,
                        mime=MIME_XLSX,
                        use_container_width=True,
                        key="dl_extra"
                    )
                except Exception as e:
                    logger.error("Error generando Excel Ordenes Extra: %s\n%s",
                                 e, traceback.format_exc())
                    st.error(f"❌ Error generando el archivo Excel: {e}")

    # ── TAB VISTA PREVIA & EXPORTAR ─────────────────────
    with tab_preview:

        df_auto   = st.session_state.get("df_auto")
        df_manual = recolectar_manuales()

        frames = []
        if df_auto is not None:
            frames.append(df_auto)
        if not df_manual.empty:
            frames.append(df_manual)

        df_total = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()

        if df_total.empty:
            st.info("⬅️ Aún no hay registros. Sube el archivo Excel o agrega registros en las pestañas anteriores.")
        else:
            df_ccr    = df_total[df_total["#"] < LIMITE_FRAVAL]
            df_fraval = df_total[df_total["#"] >= LIMITE_FRAVAL]
            n_auto    = len(df_auto) if df_auto is not None else 0
            n_manual  = len(df_manual)

            col1, col2, col3, col4 = st.columns(4)
            col1.metric("Total órdenes",                    len(df_total))
            col2.metric(f"CCR (< {LIMITE_FRAVAL})",         len(df_ccr))
            col3.metric(f"FRAVAL (≥ {LIMITE_FRAVAL})",      len(df_fraval))
            col4.metric("🤖 Auto / ✏️ Manual",              f"{n_auto} / {n_manual}")

            st.markdown("---")
            st.markdown('<div class="seccion-titulo">▸ Vista previa completa</div>', unsafe_allow_html=True)

            filas_html = ""
            for _, r in df_total.iterrows():
                p = r["Nombre Prod"]; bg = color_prod(p); fg = color_text(p)
                cliente_badge = (f'<span class="badge-ccr">CCR</span>'
                                 if r["#"] < LIMITE_FRAVAL
                                 else f'<span class="badge-fraval">FRAVAL</span>')
                origen_badge  = (f'<span class="badge-auto">Auto</span>'
                                 if r.get("origen") == "auto"
                                 else f'<span class="badge-manual">Manual</span>')
                filas_html += f"""<tr>
                    <td>{r['#']}</td>
                    <td style="background:{bg};color:{fg};">{p}</td>
                    <td>{r['Numero de Tonel']}</td>
                    <td>{r['Tipo']}</td>
                    <td>{cliente_badge}</td>
                    <td>{origen_badge}</td></tr>"""
            st.markdown(f"""<table class="preview-table">
                <tr><th>#</th><th>Producto</th><th>Tonel</th><th>Tipo</th><th>Cliente</th><th>Origen</th></tr>
                {filas_html}</table>""", unsafe_allow_html=True)

            st.markdown("---")
            st.markdown('<div class="seccion-titulo">▸ Exportar Excel</div>', unsafe_allow_html=True)

            sufijo = st.text_input(
                "Sufijo (opcional)",
                placeholder="ej: v2, v3, REVISADO — déjalo vacío si no aplica",
                key="sufijo_gen",
                help="Se agregará al final del nombre del archivo, antes de la extensión. "
                     "Útil cuando ya descargaste una versión y quieres distinguir la nueva."
            )

            archivos = {}
            try:
                if not df_ccr.empty:
                    nombre_ccr = aplicar_sufijo(
                        f"SIIC_COMPARTIMENTOS_Y_ADICIONALES_CCR_{fecha_usuario}.xlsx",
                        sufijo
                    )
                    archivos[nombre_ccr] = generar_excel(construir(df_ccr, fecha_usuario, CLIENTE_CCR))
                if not df_fraval.empty:
                    nombre_fraval = aplicar_sufijo(
                        f"SIIC_COMPARTIMENTOS_Y_ADICIONALES_FRAVAL_{fecha_usuario}.xlsx",
                        sufijo
                    )
                    archivos[nombre_fraval] = generar_excel(construir(df_fraval, fecha_usuario, CLIENTE_FRAVAL))
            except Exception as e:
                logger.error("Error generando Excel(es): %s\n%s", e, traceback.format_exc())
                st.error(f"❌ Error generando los archivos Excel: {e}")
                archivos = {}

            if len(archivos) == 1:
                nombre, datos = list(archivos.items())[0]
                st.download_button(f"📥 Descargar {nombre}", data=datos,
                                   file_name=nombre, mime=MIME_XLSX, use_container_width=True)
            elif len(archivos) == 2:
                items = list(archivos.items())
                st.markdown("**Ambos clientes detectados:**")
                col1, col2 = st.columns(2)
                with col1:
                    st.markdown(f"<small style='color:#1d4ed8;font-weight:600;'>🔵 CCR — {len(df_ccr)} orden(es)</small>",
                                unsafe_allow_html=True)
                    st.download_button(f"📥 {items[0][0]}", data=items[0][1],
                                       file_name=items[0][0], mime=MIME_XLSX, use_container_width=True)
                with col2:
                    st.markdown(f"<small style='color:#b91c1c;font-weight:600;'>🔴 FRAVAL — {len(df_fraval)} orden(es)</small>",
                                unsafe_allow_html=True)
                    st.download_button(f"📥 {items[1][0]}", data=items[1][1],
                                       file_name=items[1][0], mime=MIME_XLSX, use_container_width=True)

            st.markdown("---")
            if st.button("🧹 Limpiar capturas manuales", use_container_width=True):
                st.session_state.tn_individual = []
                st.session_state.tn_masivo     = []
                st.session_state.compartimentos_tc   = []
                st.session_state.pop("editando_ind", None)
                st.rerun()


# =========================
# LIMPIAR TODO
# =========================
st.markdown("---")
if st.button("🧹 Limpiar todo"):
    for key in list(st.session_state.keys()):
        del st.session_state[key]
    st.rerun()
